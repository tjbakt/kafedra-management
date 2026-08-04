from __future__ import annotations

from collections import defaultdict
from copy import copy
from decimal import Decimal
from typing import Iterable

from django.core.exceptions import ObjectDoesNotExist
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from apps.curriculum.models import WorkloadType
from apps.reports.exceptions import ReportDataError
from apps.reports.models import ExcelReportTemplate
from apps.reports.services.base_excel_report import (
    BaseExcelReportService,
)
from apps.staff.models import (
    StaffEmployment,
    StaffEmploymentAcademicYear,
)
from apps.workload.models import WorkloadDistribution

ZERO_HOURS = Decimal("0.00")
class TeacherWorkloadExcelService(
    BaseExcelReportService
):
    template_type = (
        ExcelReportTemplate.Type.TEACHER_WORKLOAD
    )

    TITLE_ROW = 5
    FIRST_DATA_ROW = 6
    MAX_COLUMN = 25  # Y

    CATEGORY_COLUMNS = {
        WorkloadType.ReportCategory.LECTURE: 8,
        WorkloadType.ReportCategory.PRACTICE: 9,
        WorkloadType.ReportCategory.LABORATORY: 10,

        WorkloadType.ReportCategory.COURSE_WORK_SUPERVISION: 14,
        WorkloadType.ReportCategory.COURSE_PROJECT_SUPERVISION: 14,
        WorkloadType.ReportCategory.COURSE_WORK_PROJECT_DEFENSE: 15,

        WorkloadType.ReportCategory.SCIENTIFIC_PRACTICE: 17,
        WorkloadType.ReportCategory.MASTER_DISSERTATION_SUPERVISION: 18,

        WorkloadType.ReportCategory.QUALIFICATION_PRACTICE: 22,
        WorkloadType.ReportCategory.GRADUATION_WORK_SUPERVISION: 23,
        WorkloadType.ReportCategory.RATING: 24,
    }

    WORKLOAD_SUM_COLUMNS = tuple(range(8, 26))  # H:Y
    ALL_TOTAL_COLUMNS = WORKLOAD_SUM_COLUMNS

    # ALL_TOTAL_COLUMNS = (
    #     5,  # E — студенты
    #     6,  # F — потоки
    #     7,  # G — группы
    #     *WORKLOAD_SUM_COLUMNS,
    # )

    UNSUPPORTED_CATEGORIES = {
        WorkloadType.ReportCategory.MASTER_DISSERTATION_DEFENSE,
        WorkloadType.ReportCategory.GRADUATION_WORK_DEFENSE,
        WorkloadType.ReportCategory.OTHER,
    }

    @classmethod
    def build(
        cls,
        *,
        staff_employment_id: int,
        academic_year,
    ):
        try:
            employment = (
                StaffEmployment.objects
                .select_related(
                    "staff_member",
                    "department",
                    "department__faculty",
                    "department__faculty__university",
                    "position",
                )
                .get(
                    pk=staff_employment_id,
                    is_archived=False,
                    is_active=True,
                )
            )
        except StaffEmployment.DoesNotExist as exc:
            raise ReportDataError(
                "Активное назначение преподавателя не найдено."
            ) from exc

        try:
            academic_year_record = (
                StaffEmploymentAcademicYear.objects
                .select_related(
                    "academic_year",
                    "academic_degree",
                    "academic_title",
                )
                .get(
                    staff_employment=employment,
                    academic_year=academic_year,
                    is_archived=False,
                    is_active=True,
                )
            )
        except StaffEmploymentAcademicYear.DoesNotExist as exc:
            raise ReportDataError(
                (
                    "Для преподавателя не заполнены кадровые данные "
                    f"на учебный год {academic_year}: ставка, "
                    "учёная степень и учёное звание."
                )
            ) from exc

        university_id = (
            employment
            .department
            .faculty
            .university_id
        )

        template = cls.get_template(
            university_id=university_id,
        )
        workbook, worksheet = (
            cls.load_template_workbook(template)
        )
        cls.replace_placeholders(
            worksheet,
            {
                "{FIO}": employment.staff_member.full_name,
                "{year}": str(academic_year),
                "{st}": cls.format_rate(
                    academic_year_record.rate
                ),
                "{degree}": (
                    str(academic_year_record.academic_degree)
                    if academic_year_record.academic_degree_id
                    else "-"
                ),
                "{title}": (
                    str(academic_year_record.academic_title)
                    if academic_year_record.academic_title_id
                    else "-"
                ),
            },
        )

        distributions = cls.get_distributions(
            staff_employment=employment,
            academic_year=academic_year,
        )

        rows = cls.prepare_rows(distributions)

        autumn_rows = [
            row
            for row in rows
            if row["semester_number"] % 2 == 1
        ]
        spring_rows = [
            row
            for row in rows
            if row["semester_number"] % 2 == 0
        ]

        cls.fill_report(
            worksheet=worksheet,
            autumn_rows=autumn_rows,
            spring_rows=spring_rows,
        )

        return cls.save_to_bytes(workbook)

    @staticmethod
    def format_rate(rate) -> str:
        """
        Возвращает ставку преподавателя в удобном для отображения виде.
        """

        if rate is None:
            return "-"

        value = Decimal(rate)

        text = f"{value:.2f}"

        return text.rstrip("0").rstrip(".")

    @staticmethod
    def get_distributions(
        *,
        staff_employment: StaffEmployment,
        academic_year,
    ):
        return (
            WorkloadDistribution.objects
            .filter(
                staff_employment=staff_employment,
                planned_workload__academic_year=academic_year,
                status=WorkloadDistribution.Status.APPROVED,
                is_archived=False,
                planned_workload__is_archived=False,
            )
            .select_related(
                "planned_workload",
                "planned_workload__academic_year",
                "planned_workload__academic_semester",
                "planned_workload__teaching_department",
                "planned_workload__curriculum_workload",
                "planned_workload__curriculum_workload__workload_type",
                "planned_workload__teaching_stream",
                "planned_workload__teaching_stream__curriculum_discipline",
                "planned_workload__teaching_stream__curriculum_discipline__discipline",
                "planned_workload__teaching_stream__curriculum_discipline__curriculum",
                "planned_workload__teaching_stream__curriculum_discipline__curriculum__study_program",
            )
            .prefetch_related(
                (
                    "planned_workload__teaching_stream__"
                    "stream_groups"
                ),
                (
                    "planned_workload__teaching_stream__"
                    "stream_groups__group_semester"
                ),
                (
                    "planned_workload__teaching_stream__"
                    "stream_groups__group_semester__"
                    "group_curriculum"
                ),
                (
                    "planned_workload__teaching_stream__"
                    "stream_groups__group_semester__"
                    "group_curriculum__student_group"
                ),
                (
                    "planned_workload__teaching_stream__"
                    "stream_groups__group_semester__"
                    "group_curriculum__student_group__faculty"
                ),
            )
            .order_by(
                "planned_workload__teaching_stream__curriculum_discipline__semester_number",
                "planned_workload__teaching_stream__curriculum_discipline__discipline__name_ru",
                "planned_workload__teaching_stream__code",
                "planned_workload__curriculum_workload__workload_type__sort_order",
                "pk",
            )
        )

    @classmethod
    def prepare_rows(cls, distributions: Iterable[WorkloadDistribution])-> list[dict]:
        """
        Объединяет разные виды нагрузки одной дисциплины
        в одну строку отчёта.
        """

        rows: dict[tuple, dict] = {}

        for distribution in distributions:
            planned = distribution.planned_workload
            stream = planned.teaching_stream
            curriculum_discipline = (
                stream.curriculum_discipline
            )
            discipline = curriculum_discipline.discipline
            workload_type = (
                planned.curriculum_workload.workload_type
            )

            category = workload_type.report_category
            allocated_hours = Decimal(
                distribution.allocated_hours
                or ZERO_HOURS
            )

            if (
                    category in cls.UNSUPPORTED_CATEGORIES
                    and allocated_hours > ZERO_HOURS
            ):
                raise ReportDataError(
                    f"Категория нагрузки "
                    f"«{workload_type.get_report_category_display()}» "
                    "не имеет отдельной колонки в шаблоне "
                    "отчёта преподавателя."
                )

            stream_groups = list(
                stream.stream_groups.all()
            )

            if not stream_groups:
                raise ReportDataError(
                    f"Поток «{stream.code}» не содержит "
                    "учебных групп."
                )

            group_names: list[str] = []
            faculty_names: set[str] = set()
            students_count = 0

            for stream_group in stream_groups:
                group_semester = (
                    stream_group.group_semester
                )
                student_group = (
                    group_semester.group_curriculum.student_group
                )

                group_names.append(
                    student_group.code
                )

                students_count += (
                        group_semester.students_count or 0
                )

                if student_group.faculty_id:
                    faculty_names.add(
                        student_group.faculty.name_ru
                    )

            semester_number = curriculum_discipline.semester_number

            key = (
                semester_number,
                discipline.pk,
                # tuple(sorted(group_names)),
                stream.pk,
            )

            if key not in rows:
                rows[key] = {
                    "semester_number": semester_number,
                    "discipline": discipline.name_ru,
                    "faculty": ", ".join(
                        sorted(faculty_names)
                    ),
                    "course": (
                        (semester_number + 1) // 2
                    ),
                    "students_count": students_count,
                    "streams_count": 1,
                    "groups_count": len(set(group_names)),
                    "group_names": ", ".join(
                        sorted(set(group_names))
                    ),
                    "hours": defaultdict(
                        lambda: Decimal("0.00")
                    ),
                }

            if category in cls.CATEGORY_COLUMNS:
                rows[key]["hours"][category] += (
                    allocated_hours
                )

        return sorted(
            rows.values(),
            key=lambda item: (
                item["semester_number"],
                item["discipline"],
                item["group_names"],
            ),
        )

    @classmethod
    def fill_report(
        cls,
        *,
        worksheet,
        autumn_rows: list[dict],
        spring_rows: list[dict],
    ) -> None:
        """
        Формирует структуру:

        строка 5  — Осенний семестр
        строки 6+ — данные осени
        итог осени
        заголовок весны
        данные весны
        итог весны
        итог учебного года
        """

        title_style = cls.capture_row_style(
            worksheet,
            row=cls.TITLE_ROW,
            min_column=1,
            max_column=cls.MAX_COLUMN,
        )

        data_style = cls.capture_row_style(
            worksheet,
            row=cls.FIRST_DATA_ROW,
            min_column=1,
            max_column=cls.MAX_COLUMN,
        )

        cls.remove_lower_merged_ranges(
            worksheet,
            from_row=cls.FIRST_DATA_ROW,
        )

        if worksheet.max_row >= cls.FIRST_DATA_ROW:
            worksheet.delete_rows(
                cls.FIRST_DATA_ROW,
                worksheet.max_row
                - cls.FIRST_DATA_ROW
                + 1,
            )

        worksheet.cell(
            row=cls.TITLE_ROW,
            column=1,
        ).value = "Осенний семестр"

        cls.ensure_merged_title_row(
            worksheet,
            row=cls.TITLE_ROW,
        )

        current_row = cls.FIRST_DATA_ROW

        autumn_start_row = current_row

        for number, item in enumerate(
            autumn_rows,
            start=1,
        ):
            cls.apply_row_style(
                worksheet,
                row=current_row,
                style_snapshot=data_style,
            )

            cls.write_data_row(
                worksheet=worksheet,
                row=current_row,
                number=number,
                item=item,
            )

            current_row += 1

        autumn_end_row = current_row - 1

        autumn_total_row = current_row

        cls.apply_row_style(
            worksheet,
            row=autumn_total_row,
            style_snapshot=data_style,
        )

        cls.write_total_row(
            worksheet=worksheet,
            row=autumn_total_row,
            title="Всего за осенний семестр",
            data_start_row=autumn_start_row,
            data_end_row=autumn_end_row,
        )

        current_row += 1

        spring_title_row = current_row

        cls.apply_row_style(
            worksheet,
            row=spring_title_row,
            style_snapshot=title_style,
        )

        cls.ensure_merged_title_row(
            worksheet,
            row=spring_title_row,
        )

        worksheet.cell(
            row=spring_title_row,
            column=1,
        ).value = "Весенний семестр"

        current_row += 1

        spring_start_row = current_row

        for number, item in enumerate(
            spring_rows,
            start=1,
        ):
            cls.apply_row_style(
                worksheet,
                row=current_row,
                style_snapshot=data_style,
            )

            cls.write_data_row(
                worksheet=worksheet,
                row=current_row,
                number=number,
                item=item,
            )

            current_row += 1

        spring_end_row = current_row - 1

        spring_total_row = current_row

        cls.apply_row_style(
            worksheet,
            row=spring_total_row,
            style_snapshot=data_style,
        )

        cls.write_total_row(
            worksheet=worksheet,
            row=spring_total_row,
            title="Всего за весенний семестр",
            data_start_row=spring_start_row,
            data_end_row=spring_end_row,
        )

        current_row += 1

        annual_total_row = current_row

        cls.apply_row_style(
            worksheet,
            row=annual_total_row,
            style_snapshot=data_style,
        )

        cls.write_annual_total_row(
            worksheet=worksheet,
            row=annual_total_row,
            autumn_total_row=autumn_total_row,
            spring_total_row=spring_total_row,
        )

    @classmethod
    def write_data_row(
        cls,
        *,
        worksheet,
        row: int,
        number: int,
        item: dict,
    ) -> None:
        cls.clear_row_values(
            worksheet,
            row=row,
            min_column=1,
            max_column=cls.MAX_COLUMN,
        )

        worksheet.cell(
            row=row,
            column=1,
        ).value = number

        worksheet.cell(
            row=row,
            column=2,
        ).value = item["discipline"]

        worksheet.cell(
            row=row,
            column=3,
        ).value = item["faculty"]

        worksheet.cell(
            row=row,
            column=4,
        ).value = item["course"]

        worksheet.cell(
            row=row,
            column=5,
        ).value = item["students_count"]

        worksheet.cell(
            row=row,
            column=6,
        ).value = item["streams_count"]

        worksheet.cell(
            row=row,
            column=7,
        ).value = item["groups_count"]

        for category, hours in item["hours"].items():
            column = cls.CATEGORY_COLUMNS.get(
                category
            )

            if column is None:
                continue

            worksheet.cell(
                row=row,
                column=column,
            ).value = cls.hours_or_empty(
                hours
            )
        # Y — всего по строке.
        worksheet.cell(
            row=row,
            column=25,
        ).value = f"=SUM(H{row}:X{row})"

    @classmethod
    def write_total_row(
            cls,
            *,
            worksheet,
            row: int,
            title: str,
            data_start_row: int,
            data_end_row: int,
    ) -> None:
        cls.clear_row_values(
            worksheet,
            row=row,
            min_column=1,
            max_column=cls.MAX_COLUMN,
        )

        worksheet.cell(
            row=row,
            column=2,
        ).value = title

        cls.set_sum_formulas(
            worksheet,
            total_row=row,
            data_start_row=data_start_row,
            data_end_row=data_end_row,
            sum_columns=cls.ALL_TOTAL_COLUMNS,
        )

    @classmethod
    def write_annual_total_row(
            cls,
            *,
            worksheet,
            row: int,
            autumn_total_row: int,
            spring_total_row: int,
    ) -> None:
        cls.clear_row_values(
            worksheet,
            row=row,
            min_column=1,
            max_column=cls.MAX_COLUMN,
        )

        worksheet.cell(
            row=row,
            column=2,
        ).value = "Всего за учебный год"

        for column in cls.ALL_TOTAL_COLUMNS:
            letter = get_column_letter(column)

            worksheet.cell(
                row=row,
                column=column,
            ).value = (
                f"={letter}{autumn_total_row}"
                f"+{letter}{spring_total_row}"
            )

    @staticmethod
    def capture_row_style(
            worksheet,
            *,
            row: int,
            min_column: int,
            max_column: int,
    ) -> dict:
        cells = {}

        for column in range(
                min_column,
                max_column + 1,
        ):
            cell = worksheet.cell(
                row=row,
                column=column,
            )

            if isinstance(cell, MergedCell):
                continue

            cells[column] = {
                "style": copy(cell._style),
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "number_format": cell.number_format,
                "protection": copy(cell.protection),
            }

        return {
            "height": worksheet.row_dimensions[
                row
            ].height,
            "cells": cells,
        }

    @staticmethod
    def apply_row_style(
            worksheet,
            *,
            row: int,
            style_snapshot: dict,
    ) -> None:
        worksheet.row_dimensions[row].height = (
            style_snapshot["height"]
        )

        for column, style in (
                style_snapshot["cells"].items()
        ):
            cell = worksheet.cell(
                row=row,
                column=column,
            )

            cell._style = copy(style["style"])
            cell.font = copy(style["font"])
            cell.fill = copy(style["fill"])
            cell.border = copy(style["border"])
            cell.alignment = copy(
                style["alignment"]
            )
            cell.number_format = (
                style["number_format"]
            )
            cell.protection = copy(
                style["protection"]
            )

    @classmethod
    def ensure_merged_title_row(
            cls,
            worksheet,
            *,
            row: int,
    ) -> None:
        target_range = (
            f"A{row}:"
            f"{get_column_letter(cls.MAX_COLUMN)}{row}"
        )

        for merged_range in list(
                worksheet.merged_cells.ranges
        ):
            if (
                    merged_range.min_row == row
                    and merged_range.max_row == row
            ):
                worksheet.unmerge_cells(
                    str(merged_range)
                )

        worksheet.merge_cells(target_range)

    @staticmethod
    def remove_lower_merged_ranges(
            worksheet,
            *,
            from_row: int,
    ) -> None:
        for merged_range in list(
                worksheet.merged_cells.ranges
        ):
            if merged_range.max_row >= from_row:
                worksheet.unmerge_cells(
                    str(merged_range)
                )

    @staticmethod
    def hours_or_empty(value):
        hours = Decimal(
            value or ZERO_HOURS
        )

        if hours == ZERO_HOURS:
            return None

        return float(hours)
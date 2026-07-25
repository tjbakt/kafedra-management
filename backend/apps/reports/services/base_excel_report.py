from copy import copy
from io import BytesIO
from pathlib import Path

from apps.reports.exceptions import ReportGenerationError
from django.db.models import Q
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter

from apps.reports.models import ExcelReportTemplate

from openpyxl.utils.exceptions import InvalidFileException


class BaseExcelReportService:
    template_type = None

    @classmethod
    def get_template(
        cls,
        *,
        university_id=None,
    ):
        if cls.template_type is None:
            raise ReportGenerationError(
                "Для сервиса не указан тип Excel-шаблона."
            )

        queryset = ExcelReportTemplate.objects.filter(
            template_type=cls.template_type,
            is_active=True,
            is_archived=False,
        )

        if university_id:
            template = (
                queryset
                .filter(
                    Q(university_id=university_id)
                    | Q(university__isnull=True)
                )
                .order_by(
                    "-university_id",
                    "-version",
                )
                .first()
            )
        else:
            template = (
                queryset
                .filter(university__isnull=True)
                .order_by("-version")
                .first()
            )

        if not template:
            raise ReportGenerationError(
                "Активный Excel-шаблон отчёта не найден."
            )

        if not template.file:
            raise ReportGenerationError(
                "Для шаблона не загружен Excel-файл."
            )

        return template

    @staticmethod
    def load_template_workbook(template):
        """
        Открывает загруженный Excel-шаблон.

        Исходный файл шаблона не изменяется.
        """

        try:
            template.file.open("rb")

            try:
                workbook = load_workbook(
                    filename=template.file,
                    data_only=False,
                )
            finally:
                template.file.close()

        except (
                InvalidFileException,
                OSError,
                ValueError,
        ) as exc:
            raise ReportGenerationError(
                "Не удалось открыть Excel-шаблон отчёта. "
                "Проверьте формат и целостность файла."
            ) from exc

        if template.sheet_name:
            if template.sheet_name not in workbook.sheetnames:
                workbook.close()

                raise ReportGenerationError(
                    (
                        "В Excel-шаблоне отсутствует лист "
                        f"«{template.sheet_name}»."
                    )
                )

            worksheet = workbook[
                template.sheet_name
            ]
        else:
            worksheet = workbook.active

        return workbook, worksheet

    @staticmethod
    def replace_placeholders(
        worksheet,
        replacements,
    ):
        """
        Заменяет placeholders во всех строковых ячейках листа.
        Работает и для объединённой ячейки A1:Y1,
        поскольку значение находится в её первой ячейке A1.
        """

        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue

                if not isinstance(cell.value, str):
                    continue

                value = cell.value

                for placeholder, replacement in replacements.items():
                    value = value.replace(
                        placeholder,
                        str(replacement),
                    )

                cell.value = value

    @staticmethod
    def copy_row_style(
        worksheet,
        *,
        source_row,
        target_row,
        min_column,
        max_column,
    ):
        """
        Копирует оформление строки шаблона:
        шрифт, заливку, границы, выравнивание и формат.
        """

        worksheet.row_dimensions[target_row].height = (
            worksheet.row_dimensions[source_row].height
        )

        for column in range(
            min_column,
            max_column + 1,
        ):
            source_cell = worksheet.cell(
                row=source_row,
                column=column,
            )
            target_cell = worksheet.cell(
                row=target_row,
                column=column,
            )

            if source_cell.has_style:
                target_cell._style = copy(
                    source_cell._style
                )

            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(
                source_cell.alignment
            )
            target_cell.number_format = (
                source_cell.number_format
            )
            target_cell.protection = copy(
                source_cell.protection
            )

    @staticmethod
    def clear_row_values(
        worksheet,
        *,
        row,
        min_column,
        max_column,
    ):
        for column in range(
            min_column,
            max_column + 1,
        ):
            worksheet.cell(
                row=row,
                column=column,
            ).value = None

    @staticmethod
    def set_sum_formulas(
        worksheet,
        *,
        total_row,
        data_start_row,
        data_end_row,
        sum_columns,
    ):
        for column in sum_columns:
            letter = get_column_letter(column)

            if data_end_row >= data_start_row:
                worksheet.cell(
                    row=total_row,
                    column=column,
                ).value = (
                    f"=SUM({letter}{data_start_row}:"
                    f"{letter}{data_end_row})"
                )
            else:
                worksheet.cell(
                    row=total_row,
                    column=column,
                ).value = 0

    @staticmethod
    def save_to_bytes(workbook):
        output = BytesIO()

        try:
            workbook.save(output)
        except (
                OSError,
                ValueError,
        ) as exc:
            raise ReportGenerationError(
                "Не удалось сохранить сформированный Excel-отчёт."
            ) from exc
        finally:
            workbook.close()

        output.seek(0)

        return output
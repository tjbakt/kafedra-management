import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from apps.workload.services.department_workload_service import (
    DepartmentWorkloadService,
)


class DepartmentWorkloadExportService:
    HEADERS = (
        "Кафедра",
        "Количество позиций",
        "Плановые часы",
        "Черновик",
        "Утверждено",
        "Распределено",
        "Остаток",
        "Распределение, %",
        "Статус",
    )

    STATUS_LABELS = {
        "incomplete": "Не завершено",
        "complete": "Завершено",
        "exceeded": "Превышение",
    }

    @classmethod
    def export(
        cls,
        *,
        academic_year,
        academic_semester_id=None,
        department_id=None,
    ) -> tuple[bytes, str]:
        summary = DepartmentWorkloadService.get_summary(
            academic_year=academic_year,
            academic_semester_id=academic_semester_id,
            department_id=department_id,
        )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Сводка по кафедрам"

        cls._write_title(
            worksheet=worksheet,
            academic_year=academic_year,
        )
        cls._write_headers(worksheet)
        cls._write_rows(
            worksheet=worksheet,
            summary=summary,
        )
        cls._write_totals(
            worksheet=worksheet,
            summary=summary,
        )
        cls._configure_worksheet(worksheet)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        filename = cls._build_filename(
            academic_year_name=academic_year.name,
        )

        return buffer.getvalue(), filename

    @classmethod
    def _write_title(
        cls,
        *,
        worksheet,
        academic_year,
    ):
        worksheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=len(cls.HEADERS),
        )

        cell = worksheet.cell(
            row=1,
            column=1,
            value=(
                "Сводка распределения нагрузки "
                f"за {academic_year.name} учебный год"
            ),
        )
        cell.font = Font(
            bold=True,
            size=14,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        worksheet.row_dimensions[1].height = 24

    @classmethod
    def _write_headers(cls, worksheet):
        header_row = 3

        for column, header in enumerate(
            cls.HEADERS,
            start=1,
        ):
            cell = worksheet.cell(
                row=header_row,
                column=column,
                value=header,
            )
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

    @classmethod
    def _write_rows(
        cls,
        *,
        worksheet,
        summary,
    ):
        start_row = 4

        for row_number, item in enumerate(
            summary,
            start=start_row,
        ):
            values = (
                item["department_name"],
                item["planned_positions"],
                item["planned_hours"],
                item["draft_hours"],
                item["approved_hours"],
                item["distributed_hours"],
                item["remaining_hours"],
                item["distribution_percent"],
                cls.STATUS_LABELS.get(
                    item["distribution_status"],
                    item["distribution_status"],
                ),
            )

            for column, value in enumerate(
                values,
                start=1,
            ):
                cell = worksheet.cell(
                    row=row_number,
                    column=column,
                    value=value,
                )

                if 3 <= column <= 8:
                    cell.number_format = "0.00"

                if column == 9:
                    cell.alignment = Alignment(
                        horizontal="center",
                    )

    @classmethod
    def _write_totals(
        cls,
        *,
        worksheet,
        summary,
    ):
        total_row = len(summary) + 5

        worksheet.cell(
            row=total_row,
            column=1,
            value="Итого",
        ).font = Font(bold=True)

        totals = {
            2: sum(
                item["planned_positions"]
                for item in summary
            ),
            3: sum(
                item["planned_hours"]
                for item in summary
            ),
            4: sum(
                item["draft_hours"]
                for item in summary
            ),
            5: sum(
                item["approved_hours"]
                for item in summary
            ),
            6: sum(
                item["distributed_hours"]
                for item in summary
            ),
            7: sum(
                item["remaining_hours"]
                for item in summary
            ),
        }

        for column, value in totals.items():
            cell = worksheet.cell(
                row=total_row,
                column=column,
                value=value,
            )
            cell.font = Font(bold=True)

            if column >= 3:
                cell.number_format = "0.00"

        planned_hours = totals[3]
        distributed_hours = totals[6]

        if planned_hours:
            total_percent = (
                distributed_hours
                / planned_hours
                * 100
            )
        else:
            total_percent = 0

        percent_cell = worksheet.cell(
            row=total_row,
            column=8,
            value=total_percent,
        )
        percent_cell.font = Font(bold=True)
        percent_cell.number_format = "0.00"

    @classmethod
    def _configure_worksheet(cls, worksheet):
        worksheet.freeze_panes = "A4"
        worksheet.auto_filter.ref = (
            f"A3:I{worksheet.max_row}"
        )

        widths = {
            1: 42,
            2: 18,
            3: 16,
            4: 14,
            5: 14,
            6: 16,
            7: 14,
            8: 18,
            9: 18,
        }

        for column, width in widths.items():
            worksheet.column_dimensions[
                get_column_letter(column)
            ].width = width

    @staticmethod
    def _build_filename(
        *,
        academic_year_name,
    ):
        safe_year_name = re.sub(
            r"[^0-9A-Za-zА-Яа-я_-]+",
            "_",
            academic_year_name,
        ).strip("_")

        return (
            "department_workload_summary_"
            f"{safe_year_name}.xlsx"
        )
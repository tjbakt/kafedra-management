import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from apps.workload.services.workload_dashboard_service import (
    WorkloadDashboardService,
)


class WorkloadDashboardExportService:
    CONTENT_TYPE = (
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    )

    THIN_BORDER = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    @classmethod
    def export(
        cls,
        *,
        academic_year,
        department_id=None,
        allowed_department_ids=None,
        allowed_staff_member_ids=None,
    ) -> tuple[bytes, str]:
        dashboard = WorkloadDashboardService.get_dashboard(
            academic_year=academic_year,
            department_id=department_id,
            allowed_department_ids=allowed_department_ids,
            allowed_staff_member_ids=allowed_staff_member_ids,
        )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Сводный дашборд"

        cls._write_title(
            worksheet=worksheet,
            dashboard=dashboard,
        )
        cls._write_workload_section(
            worksheet=worksheet,
            dashboard=dashboard,
        )
        cls._write_teacher_section(
            worksheet=worksheet,
            dashboard=dashboard,
        )
        cls._write_department_section(
            worksheet=worksheet,
            dashboard=dashboard,
        )
        cls._configure_worksheet(worksheet)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        filename = cls._build_filename(
            academic_year_name=academic_year.name,
            department_id=department_id,
        )

        return buffer.getvalue(), filename

    @classmethod
    def _write_title(
        cls,
        *,
        worksheet,
        dashboard,
    ):
        worksheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=3,
        )

        title = (
            "Сводный отчёт по распределению нагрузки "
            f"за {dashboard['academic_year_name']} учебный год"
        )

        if dashboard["department"] is not None:
            title += (
                f", кафедра ID: {dashboard['department']}"
            )

        cell = worksheet.cell(
            row=1,
            column=1,
            value=title,
        )
        cell.font = Font(
            bold=True,
            size=14,
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

        worksheet.row_dimensions[1].height = 36

    @classmethod
    def _write_workload_section(
        cls,
        *,
        worksheet,
        dashboard,
    ):
        workload = dashboard["workload"]

        rows = (
            (
                "Количество плановых позиций",
                workload["planned_positions"],
                "позиций",
            ),
            (
                "Плановые часы",
                workload["planned_hours"],
                "часов",
            ),
            (
                "Черновые распределения",
                workload["draft_hours"],
                "часов",
            ),
            (
                "Утверждённые распределения",
                workload["approved_hours"],
                "часов",
            ),
            (
                "Всего распределено",
                workload["distributed_hours"],
                "часов",
            ),
            (
                "Остаток",
                workload["remaining_hours"],
                "часов",
            ),
            (
                "Процент распределения",
                workload["distribution_percent"],
                "%",
            ),
        )

        cls._write_section(
            worksheet=worksheet,
            start_row=3,
            title="Общие показатели нагрузки",
            rows=rows,
            decimal_rows={2, 3, 4, 5, 6, 7},
        )

    @classmethod
    def _write_teacher_section(
        cls,
        *,
        worksheet,
        dashboard,
    ):
        teachers = dashboard["teachers"]

        rows = (
            (
                "Всего преподавателей",
                teachers["total"],
                "человек",
            ),
            (
                "С найденной нормой",
                teachers["with_norm"],
                "человек",
            ),
            (
                "Без найденной нормы",
                teachers["without_norm"],
                "человек",
            ),
            (
                "С недогрузкой",
                teachers["underloaded"],
                "человек",
            ),
            (
                "Норма выполнена",
                teachers["balanced"],
                "человек",
            ),
            (
                "С перегрузкой",
                teachers["overloaded"],
                "человек",
            ),
            (
                "Рекомендуемые часы",
                teachers["recommended_hours"],
                "часов",
            ),
            (
                "Распределённые часы",
                teachers["distributed_hours"],
                "часов",
            ),
        )

        cls._write_section(
            worksheet=worksheet,
            start_row=13,
            title="Показатели преподавателей",
            rows=rows,
            decimal_rows={7, 8},
        )

    @classmethod
    def _write_department_section(
        cls,
        *,
        worksheet,
        dashboard,
    ):
        departments = dashboard["departments"]

        rows = (
            (
                "Всего кафедр",
                departments["total"],
                "кафедр",
            ),
            (
                "Распределение не завершено",
                departments["incomplete"],
                "кафедр",
            ),
            (
                "Распределение завершено",
                departments["complete"],
                "кафедр",
            ),
            (
                "План превышен",
                departments["exceeded"],
                "кафедр",
            ),
        )

        cls._write_section(
            worksheet=worksheet,
            start_row=24,
            title="Состояние кафедр",
            rows=rows,
            decimal_rows=set(),
        )

    @classmethod
    def _write_section(
        cls,
        *,
        worksheet,
        start_row,
        title,
        rows,
        decimal_rows,
    ):
        worksheet.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=start_row,
            end_column=3,
        )

        title_cell = worksheet.cell(
            row=start_row,
            column=1,
            value=title,
        )
        title_cell.font = Font(
            bold=True,
            size=12,
        )
        title_cell.alignment = Alignment(
            horizontal="left",
            vertical="center",
        )

        header_row = start_row + 1

        headers = (
            "Показатель",
            "Значение",
            "Единица измерения",
        )

        for column, header in enumerate(
            headers,
            start=1,
        ):
            cell = worksheet.cell(
                row=header_row,
                column=column,
                value=header,
            )
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )
            cell.border = cls.THIN_BORDER

        for offset, row_values in enumerate(
            rows,
            start=1,
        ):
            current_row = header_row + offset

            for column, value in enumerate(
                row_values,
                start=1,
            ):
                cell = worksheet.cell(
                    row=current_row,
                    column=column,
                    value=cls._excel_value(value),
                )
                cell.border = cls.THIN_BORDER
                cell.alignment = Alignment(
                    vertical="center",
                    horizontal=(
                        "left"
                        if column == 1
                        else "center"
                    ),
                )

            if offset in decimal_rows:
                worksheet.cell(
                    row=current_row,
                    column=2,
                ).number_format = "0.00"

    @staticmethod
    def _excel_value(value):
        if value is None:
            return None

        try:
            return float(value)
        except (TypeError, ValueError):
            return value

    @staticmethod
    def _configure_worksheet(worksheet):
        widths = {
            1: 38,
            2: 20,
            3: 24,
        }

        for column, width in widths.items():
            worksheet.column_dimensions[
                get_column_letter(column)
            ].width = width

        worksheet.freeze_panes = "A3"
        worksheet.sheet_view.showGridLines = False

    @staticmethod
    def _build_filename(
        *,
        academic_year_name,
        department_id=None,
    ):
        safe_year_name = re.sub(
            r"[^0-9A-Za-zА-Яа-я_-]+",
            "_",
            academic_year_name,
        ).strip("_")

        filename = (
            f"workload_dashboard_{safe_year_name}"
        )

        if department_id:
            filename += (
                f"_department_{department_id}"
            )

        return f"{filename}.xlsx"
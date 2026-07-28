import re
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from apps.workload.services.teacher_workload_service import (
    TeacherWorkloadService,
)


class TeacherWorkloadExportService:
    HEADERS = (
        "Табельный номер",
        "Преподаватель",
        "Кафедра",
        "Должность",
        "Ставка",
        "Учёная степень",
        "Учёное звание",
        "Рекомендуемая норма",
        "Распределено",
        "Остаток",
        "Отклонение",
        "Выполнение, %",
        "Статус",
    )

    STATUS_LABELS = {
        "underloaded": "Недогрузка",
        "balanced": "Норма выполнена",
        "overloaded": "Перегрузка",
        "norm_missing": "Норма не найдена",
    }

    BOOLEAN_LABELS = {
        True: "Да",
        False: "Нет",
    }

    @classmethod
    def export(
        cls,
        *,
        academic_year,
        staff_member_id=None,
        department_id=None,
    ) -> tuple[bytes, str]:
        summary = TeacherWorkloadService.get_summary(
            academic_year=academic_year,
            staff_member_id=staff_member_id,
            department_id=department_id,
        )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Нагрузка преподавателей"

        cls._write_title(
            worksheet=worksheet,
            academic_year=academic_year,
        )
        cls._write_headers(worksheet)
        cls._write_rows(
            worksheet=worksheet,
            summary=summary,
        )
        cls._write_totals(
            worksheet=worksheet,
            summary=summary,
        )
        cls._configure_worksheet(worksheet)

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        filename = cls._build_filename(
            academic_year_name=academic_year.name,
        )

        return buffer.getvalue(), filename

    @classmethod
    def _write_title(
        cls,
        *,
        worksheet,
        academic_year,
    ):
        worksheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=1,
            end_column=len(cls.HEADERS),
        )

        title_cell = worksheet.cell(
            row=1,
            column=1,
            value=(
                "Сводка нагрузки преподавателей "
                f"за {academic_year.name} учебный год"
            ),
        )
        title_cell.font = Font(
            bold=True,
            size=14,
        )
        title_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        worksheet.row_dimensions[1].height = 24

    @classmethod
    def _write_headers(cls, worksheet):
        header_row = 3

        for column, header in enumerate(
            cls.HEADERS,
            start=1,
        ):
            cell = worksheet.cell(
                row=header_row,
                column=column,
                value=header,
            )
            cell.font = Font(bold=True)
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True,
            )

        worksheet.row_dimensions[header_row].height = 34

    @classmethod
    def _write_rows(
        cls,
        *,
        worksheet,
        summary,
    ):
        for row_number, item in enumerate(
            summary,
            start=4,
        ):
            values = (
                item["personnel_number"],
                item["teacher_name"],
                item["department_name"],
                item["position_name"],
                cls._excel_number(item["employment_rate"]),
                cls.BOOLEAN_LABELS[
                    item["has_academic_degree"]
                ],
                cls.BOOLEAN_LABELS[
                    item["has_academic_title"]
                ],
                cls._excel_number(
                    item["recommended_hours"]
                ),
                cls._excel_number(
                    item["distributed_hours"]
                ),
                cls._excel_number(
                    item["remaining_hours"]
                ),
                cls._excel_number(
                    item["difference_hours"]
                ),
                cls._excel_number(
                    item["load_percent"]
                ),
                cls.STATUS_LABELS.get(
                    item["load_status"],
                    item["load_status"],
                ),
            )

            for column, value in enumerate(
                values,
                start=1,
            ):
                cell = worksheet.cell(
                    row=row_number,
                    column=column,
                    value=value,
                )

                if column in (5, 8, 9, 10, 11, 12):
                    cell.number_format = "0.00"

                if column in (5, 6, 7, 12, 13):
                    cell.alignment = Alignment(
                        horizontal="center",
                        vertical="center",
                    )

    @classmethod
    def _write_totals(
        cls,
        *,
        worksheet,
        summary,
    ):
        total_row = len(summary) + 5

        worksheet.cell(
            row=total_row,
            column=1,
            value="Итого",
        ).font = Font(bold=True)

        worksheet.merge_cells(
            start_row=total_row,
            start_column=1,
            end_row=total_row,
            end_column=7,
        )

        recommended_values = [
            item["recommended_hours"]
            for item in summary
            if item["recommended_hours"] is not None
        ]
        distributed_values = [
            item["distributed_hours"]
            for item in summary
        ]
        remaining_values = [
            item["remaining_hours"]
            for item in summary
            if item["remaining_hours"] is not None
        ]
        difference_values = [
            item["difference_hours"]
            for item in summary
            if item["difference_hours"] is not None
        ]

        total_recommended = sum(
            recommended_values,
            Decimal("0.00"),
        )
        total_distributed = sum(
            distributed_values,
            Decimal("0.00"),
        )
        total_remaining = sum(
            remaining_values,
            Decimal("0.00"),
        )
        total_difference = sum(
            difference_values,
            Decimal("0.00"),
        )

        totals = {
            8: total_recommended,
            9: total_distributed,
            10: total_remaining,
            11: total_difference,
        }

        for column, value in totals.items():
            cell = worksheet.cell(
                row=total_row,
                column=column,
                value=cls._excel_number(value),
            )
            cell.font = Font(bold=True)
            cell.number_format = "0.00"

        if total_recommended > Decimal("0.00"):
            total_percent = (
                total_distributed
                / total_recommended
                * Decimal("100.00")
            ).quantize(Decimal("0.01"))
        else:
            total_percent = None

        percent_cell = worksheet.cell(
            row=total_row,
            column=12,
            value=cls._excel_number(total_percent),
        )
        percent_cell.font = Font(bold=True)
        percent_cell.number_format = "0.00"

        worksheet.cell(
            row=total_row,
            column=13,
            value=f"Преподавателей: {len(summary)}",
        ).font = Font(bold=True)

    @classmethod
    def _configure_worksheet(cls, worksheet):
        worksheet.freeze_panes = "A4"

        last_data_row = max(
            worksheet.max_row,
            3,
        )
        worksheet.auto_filter.ref = (
            f"A3:M{last_data_row}"
        )

        widths = {
            1: 18,
            2: 34,
            3: 34,
            4: 26,
            5: 12,
            6: 17,
            7: 17,
            8: 21,
            9: 16,
            10: 16,
            11: 16,
            12: 17,
            13: 20,
        }

        for column, width in widths.items():
            worksheet.column_dimensions[
                get_column_letter(column)
            ].width = width

    @staticmethod
    def _excel_number(value):
        if value is None:
            return None

        if isinstance(value, Decimal):
            return float(value)

        return value

    @staticmethod
    def _build_filename(
        *,
        academic_year_name,
    ):
        safe_year_name = re.sub(
            r"[^0-9A-Za-zА-Яа-я_-]+",
            "_",
            academic_year_name,
        ).strip("_")

        return (
            "teacher_workload_summary_"
            f"{safe_year_name}.xlsx"
        )
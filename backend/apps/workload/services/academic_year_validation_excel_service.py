import json
from io import BytesIO

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)


from apps.workload.services.academic_year_validation_service import (
    AcademicYearWorkloadValidationService,
)


class AcademicYearValidationExcelService:
    """
    Формирует Excel-отчёт по результатам проверки
    учебной нагрузки за учебный год.

    Бизнес-правила проверки не реализуются здесь.
    Сервис принимает готовый результат
    AcademicYearWorkloadValidationService.validate().
    """

    MIME_TYPE = (
        "application/vnd.openxmlformats-officedocument."
        "spreadsheetml.sheet"
    )

    HEADER_FILL = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )
    SUBHEADER_FILL = PatternFill(
        fill_type="solid",
        fgColor="D9EAF7",
    )
    ERROR_FILL = PatternFill(
        fill_type="solid",
        fgColor="F4CCCC",
    )
    WARNING_FILL = PatternFill(
        fill_type="solid",
        fgColor="FFF2CC",
    )
    SUCCESS_FILL = PatternFill(
        fill_type="solid",
        fgColor="D9EAD3",
    )

    WHITE_FONT = Font(
        color="FFFFFF",
        bold=True,
    )
    BOLD_FONT = Font(
        bold=True,
    )
    TITLE_FONT = Font(
        bold=True,
        size=16,
    )

    THIN_BORDER = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    ISSUE_TYPE_LABELS = {
        AcademicYearWorkloadValidationService
        .IssueType
        .PLANNED_WITHOUT_DISTRIBUTIONS: (
            "Плановая нагрузка не распределена"
        ),
        AcademicYearWorkloadValidationService
        .IssueType
        .PLANNED_PARTIALLY_DISTRIBUTED: (
            "Плановая нагрузка распределена частично"
        ),
        AcademicYearWorkloadValidationService
        .IssueType
        .PLANNED_HOURS_EXCEEDED: (
            "Превышение плановых часов"
        ),
        AcademicYearWorkloadValidationService
        .IssueType
        .PLANNED_STATUS_MISMATCH: (
            "Несоответствие статуса плановой нагрузки"
        ),
        AcademicYearWorkloadValidationService
        .IssueType
        .DISTRIBUTION_DRAFT: (
            "Распределение находится в черновике"
        ),
        AcademicYearWorkloadValidationService
        .IssueType
        .APPROVAL_DATA_MISSING: (
            "Отсутствуют данные утверждения"
        ),
        AcademicYearWorkloadValidationService
        .IssueType
        .APPROVAL_DATA_UNEXPECTED: (
            "Лишние данные утверждения"
        ),
        AcademicYearWorkloadValidationService
        .IssueType
        .EMPLOYMENT_ARCHIVED: (
            "Архивное трудовое назначение"
        ),
        AcademicYearWorkloadValidationService
        .IssueType
        .EMPLOYMENT_INACTIVE: (
            "Неактивное трудовое назначение"
        ),
        AcademicYearWorkloadValidationService
        .IssueType
        .NON_TEACHING_POSITION: (
            "Непреподавательская должность"
        ),
        AcademicYearWorkloadValidationService
        .IssueType
        .EMPLOYMENT_DEPARTMENT_MISMATCH: (
            "Несовпадение кафедр"
        ),
        AcademicYearWorkloadValidationService
        .IssueType
        .YEAR_STAFF_RECORD_MISSING: (
            "Отсутствует годовая кадровая запись"
        ),
        AcademicYearWorkloadValidationService
        .IssueType
        .WORKLOAD_NORM_MISSING: (
            "Отсутствует норма нагрузки"
        ),
        AcademicYearWorkloadValidationService
        .IssueType
        .TEACHER_OVERLOADED: (
            "Превышена рекомендуемая норма"
        ),
    }

    SEVERITY_LABELS = {
        AcademicYearWorkloadValidationService
        .Severity.ERROR: "Ошибка",
        AcademicYearWorkloadValidationService
        .Severity.WARNING: "Предупреждение",
    }

    @classmethod
    def build(
        cls,
        *,
        validation_result,
        generated_by=None,
    ) -> BytesIO:
        workbook = Workbook()

        summary_sheet = workbook.active
        summary_sheet.title = "Сводка"

        issues_sheet = workbook.create_sheet(
            title="Проблемы"
        )
        types_sheet = workbook.create_sheet(
            title="Типы проблем"
        )

        generated_at = timezone.localtime(
            timezone.now()
        )

        cls._build_summary_sheet(
            worksheet=summary_sheet,
            validation_result=validation_result,
            generated_at=generated_at,
            generated_by=generated_by,
        )

        cls._build_issues_sheet(
            worksheet=issues_sheet,
            issues=validation_result["issues"],
        )

        cls._build_issue_types_sheet(
            worksheet=types_sheet,
            issues_by_type=(
                validation_result["summary"][
                    "issues_by_type"
                ]
            ),
        )

        cls._set_workbook_properties(
            workbook=workbook,
            validation_result=validation_result,
            generated_at=generated_at,
        )

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        return output

    @classmethod
    def build_filename(
        cls,
        *,
        validation_result,
    ) -> str:
        academic_year_name = cls._safe_filename_part(
            validation_result["academic_year_name"]
        )

        date_suffix = timezone.localdate().strftime(
            "%Y-%m-%d"
        )

        return (
            "academic-year-validation-"
            f"{academic_year_name}-"
            f"{date_suffix}.xlsx"
        )

    @classmethod
    def _build_summary_sheet(
        cls,
        *,
        worksheet,
        validation_result,
        generated_at,
        generated_by,
    ):
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A9"

        worksheet.merge_cells("A1:D1")
        title_cell = worksheet["A1"]
        title_cell.value = (
            "Проверка целостности учебной нагрузки"
        )
        title_cell.font = cls.TITLE_FONT
        title_cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        worksheet.row_dimensions[1].height = 28

        user_name = cls._user_display_name(
            generated_by
        )

        metadata = [
            (
                "Учебный год",
                validation_result[
                    "academic_year_name"
                ],
            ),
            (
                "ID учебного года",
                validation_result["academic_year"],
            ),
            (
                "Кафедры",
                cls._department_filter_text(
                    validation_result[
                        "department_ids"
                    ]
                ),
            ),
            (
                "Дата формирования",
                generated_at.strftime(
                    "%d.%m.%Y %H:%M"
                ),
            ),
            (
                "Сформировал",
                user_name,
            ),
            (
                "Результат проверки",
                (
                    "Ошибок не обнаружено"
                    if validation_result["is_valid"]
                    else "Обнаружены ошибки"
                ),
            ),
        ]

        start_row = 3

        for index, (
            label,
            value,
        ) in enumerate(
            metadata,
            start=start_row,
        ):
            label_cell = worksheet.cell(
                row=index,
                column=1,
                value=label,
            )
            value_cell = worksheet.cell(
                row=index,
                column=2,
                value=value,
            )

            label_cell.font = cls.BOLD_FONT
            label_cell.fill = cls.SUBHEADER_FILL
            label_cell.border = cls.THIN_BORDER

            value_cell.border = cls.THIN_BORDER
            value_cell.alignment = Alignment(
                wrap_text=True,
                vertical="top",
            )

        result_cell = worksheet.cell(
            row=start_row + len(metadata) - 1,
            column=2,
        )
        result_cell.fill = (
            cls.SUCCESS_FILL
            if validation_result["is_valid"]
            else cls.ERROR_FILL
        )
        result_cell.font = cls.BOLD_FONT

        summary = validation_result["summary"]

        summary_start_row = 10

        worksheet.merge_cells(
            start_row=summary_start_row,
            start_column=1,
            end_row=summary_start_row,
            end_column=4,
        )

        summary_title = worksheet.cell(
            row=summary_start_row,
            column=1,
            value="Сводные показатели",
        )
        summary_title.fill = cls.HEADER_FILL
        summary_title.font = cls.WHITE_FONT
        summary_title.alignment = Alignment(
            horizontal="center",
        )

        headers = (
            "Показатель",
            "Значение",
            "Показатель",
            "Значение",
        )

        header_row = summary_start_row + 1

        for column, header in enumerate(
            headers,
            start=1,
        ):
            cell = worksheet.cell(
                row=header_row,
                column=column,
                value=header,
            )
            cls._style_header_cell(cell)

        rows = [
            (
                "Плановых позиций",
                summary[
                    "planned_workloads_count"
                ],
                "Распределений",
                summary[
                    "distributions_count"
                ],
            ),
            (
                "Годовых кадровых записей",
                summary[
                    "year_staff_records_count"
                ],
                "Всего проблем",
                summary["issues_count"],
            ),
            (
                "Ошибок",
                summary["errors_count"],
                "Предупреждений",
                summary["warnings_count"],
            ),
        ]

        for row_number, values in enumerate(
            rows,
            start=header_row + 1,
        ):
            for column, value in enumerate(
                values,
                start=1,
            ):
                cell = worksheet.cell(
                    row=row_number,
                    column=column,
                    value=value,
                )
                cell.border = cls.THIN_BORDER
                cell.alignment = Alignment(
                    vertical="center",
                    wrap_text=True,
                )

                if column in (1, 3):
                    cell.font = cls.BOLD_FONT

        errors_cell = worksheet.cell(
            row=header_row + 3,
            column=2,
        )
        warnings_cell = worksheet.cell(
            row=header_row + 3,
            column=4,
        )

        if errors_cell.value:
            errors_cell.fill = cls.ERROR_FILL

        if warnings_cell.value:
            warnings_cell.fill = cls.WARNING_FILL

        worksheet.column_dimensions["A"].width = 32
        worksheet.column_dimensions["B"].width = 28
        worksheet.column_dimensions["C"].width = 32
        worksheet.column_dimensions["D"].width = 20

        worksheet.auto_filter.ref = (
            f"A{header_row}:D{header_row + len(rows)}"
        )

    @classmethod
    def _build_issues_sheet(
        cls,
        *,
        worksheet,
        issues,
    ):
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A2"

        headers = (
            "№",
            "Уровень",
            "Тип проблемы",
            "Сообщение",
            "Кафедра",
            "Преподаватель",
            "Поток",
            "Дисциплина",
            "Вид нагрузки",
            "ID плановой нагрузки",
            "ID распределения",
            "ID назначения",
            "Подробности",
        )

        for column, header in enumerate(
            headers,
            start=1,
        ):
            cell = worksheet.cell(
                row=1,
                column=column,
                value=header,
            )
            cls._style_header_cell(cell)

        if not issues:
            worksheet.merge_cells(
                start_row=2,
                start_column=1,
                end_row=2,
                end_column=len(headers),
            )

            cell = worksheet.cell(
                row=2,
                column=1,
                value="Проблемы не обнаружены.",
            )
            cell.fill = cls.SUCCESS_FILL
            cell.font = cls.BOLD_FONT
            cell.alignment = Alignment(
                horizontal="center",
            )

        for index, issue in enumerate(
            issues,
            start=1,
        ):
            row_number = index + 1

            values = (
                index,
                cls.SEVERITY_LABELS.get(
                    issue["severity"],
                    issue["severity"],
                ),
                cls.ISSUE_TYPE_LABELS.get(
                    issue["issue_type"],
                    issue["issue_type"],
                ),
                issue["message"],
                issue.get("department_name"),
                issue.get("teacher_name"),
                issue.get("stream_code"),
                issue.get("discipline_name"),
                issue.get("workload_type_name"),
                issue.get("planned_workload_id"),
                issue.get("distribution_id"),
                issue.get("staff_employment_id"),
                cls._serialize_details(
                    issue.get("details")
                ),
            )

            for column, value in enumerate(
                values,
                start=1,
            ):
                cell = worksheet.cell(
                    row=row_number,
                    column=column,
                    value=value,
                )
                cell.border = cls.THIN_BORDER
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

            severity_cell = worksheet.cell(
                row=row_number,
                column=2,
            )

            if (
                issue["severity"]
                == AcademicYearWorkloadValidationService
                .Severity.ERROR
            ):
                severity_cell.fill = cls.ERROR_FILL
            else:
                severity_cell.fill = cls.WARNING_FILL

            severity_cell.font = cls.BOLD_FONT

        worksheet.auto_filter.ref = (
            f"A1:M{max(1, len(issues) + 1)}"
        )

        widths = {
            "A": 7,
            "B": 18,
            "C": 38,
            "D": 55,
            "E": 35,
            "F": 35,
            "G": 18,
            "H": 35,
            "I": 28,
            "J": 20,
            "K": 18,
            "L": 18,
            "M": 60,
        }

        for column, width in widths.items():
            worksheet.column_dimensions[
                column
            ].width = width

        worksheet.page_setup.orientation = "landscape"
        worksheet.page_setup.fitToWidth = 1
        worksheet.sheet_properties.pageSetUpPr.fitToPage = (
            True
        )

    @classmethod
    def _build_issue_types_sheet(
        cls,
        *,
        worksheet,
        issues_by_type,
    ):
        worksheet.sheet_view.showGridLines = False
        worksheet.freeze_panes = "A2"

        headers = (
            "Код",
            "Наименование",
            "Количество",
        )

        for column, header in enumerate(
            headers,
            start=1,
        ):
            cell = worksheet.cell(
                row=1,
                column=column,
                value=header,
            )
            cls._style_header_cell(cell)

        sorted_items = sorted(
            issues_by_type.items(),
            key=lambda item: (
                -item[1],
                item[0],
            ),
        )

        if not sorted_items:
            worksheet.merge_cells(
                start_row=2,
                start_column=1,
                end_row=2,
                end_column=3,
            )

            cell = worksheet.cell(
                row=2,
                column=1,
                value="Проблемы не обнаружены.",
            )
            cell.fill = cls.SUCCESS_FILL
            cell.font = cls.BOLD_FONT
            cell.alignment = Alignment(
                horizontal="center",
            )

        for row_number, (
            issue_type,
            count,
        ) in enumerate(
            sorted_items,
            start=2,
        ):
            values = (
                issue_type,
                cls.ISSUE_TYPE_LABELS.get(
                    issue_type,
                    issue_type,
                ),
                count,
            )

            for column, value in enumerate(
                values,
                start=1,
            ):
                cell = worksheet.cell(
                    row=row_number,
                    column=column,
                    value=value,
                )
                cell.border = cls.THIN_BORDER
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        worksheet.auto_filter.ref = (
            f"A1:C{max(1, len(sorted_items) + 1)}"
        )

        worksheet.column_dimensions["A"].width = 42
        worksheet.column_dimensions["B"].width = 55
        worksheet.column_dimensions["C"].width = 16

    @classmethod
    def _style_header_cell(
        cls,
        cell,
    ):
        cell.fill = cls.HEADER_FILL
        cell.font = cls.WHITE_FONT
        cell.border = cls.THIN_BORDER
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )

    @staticmethod
    def _serialize_details(details) -> str:
        if not details:
            return ""

        return json.dumps(
            details,
            ensure_ascii=False,
            sort_keys=True,
            indent=2,
            default=str,
        )

    @staticmethod
    def _department_filter_text(
        department_ids,
    ) -> str:
        if not department_ids:
            return "Все доступные кафедры"

        return ", ".join(
            str(department_id)
            for department_id in department_ids
        )

    @staticmethod
    def _user_display_name(user) -> str:
        if user is None:
            return ""

        get_full_name = getattr(
            user,
            "get_full_name",
            None,
        )

        if callable(get_full_name):
            full_name = get_full_name().strip()

            if full_name:
                return full_name

        return str(user)

    @staticmethod
    def _safe_filename_part(value) -> str:
        normalized = str(value or "").strip()

        for character in (
            "/",
            "\\",
            ":",
            "*",
            "?",
            '"',
            "<",
            ">",
            "|",
            " ",
        ):
            normalized = normalized.replace(
                character,
                "-",
            )

        while "--" in normalized:
            normalized = normalized.replace(
                "--",
                "-",
            )

        return normalized.strip("-") or "academic-year"

    @staticmethod
    def _set_workbook_properties(
        *,
        workbook,
        validation_result,
        generated_at,
    ):
        workbook.properties.title = (
            "Проверка целостности учебной нагрузки"
        )
        workbook.properties.subject = (
            validation_result["academic_year_name"]
        )
        workbook.properties.creator = (
            "Kafedra Management"
        )
        workbook.properties.description = (
            "Отчёт по результатам проверки "
            "целостности учебной нагрузки"
        )
        workbook.properties.created = (
            generated_at.replace(tzinfo=None)
        )
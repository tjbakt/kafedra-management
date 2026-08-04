from decimal import Decimal

from django.test import TestCase
from openpyxl import load_workbook

from apps.curriculum.models import (
    WorkloadType,
)
from apps.reports.exceptions import (
    ReportDataError,
)
from apps.reports.models import (
    ExcelReportTemplate,
)
from apps.reports.services.department_workload_excel import (
    DepartmentWorkloadExcelService,
)
from apps.reports.services.teacher_workload_excel import (
    TeacherWorkloadExcelService,
)
from tests.factories import (
    ExcelReportTemplateFactory,
    PlannedWorkloadFactory,
    StaffEmploymentAcademicYearFactory,
    TeachingStreamGroupFactory,
    WorkloadDistributionFactory,
    WorkloadTypeFactory,
    create_xlsx_file,
)


class TeacherWorkloadExcelServiceTests(
    TestCase
):
    def create_distribution(self):
        distribution = (
            WorkloadDistributionFactory(
                status="approved",
                allocated_hours=(
                    Decimal("30.00")
                ),
                planned_workload__curriculum_workload__workload_type=(
                    WorkloadTypeFactory(
                        report_category=(
                            WorkloadType
                            .ReportCategory
                            .LECTURE
                        )
                    )
                ),
            )
        )

        TeachingStreamGroupFactory(
            teaching_stream=(
                distribution
                .planned_workload
                .teaching_stream
            )
        )

        return distribution

    def test_format_rate(self):
        self.assertEqual(
            TeacherWorkloadExcelService
            .format_rate(
                Decimal("1.00")
            ),
            "1",
        )
        self.assertEqual(
            TeacherWorkloadExcelService
            .format_rate(
                Decimal("0.50")
            ),
            "0.5",
        )
        self.assertEqual(
            TeacherWorkloadExcelService
            .format_rate(None),
            "-",
        )

    def test_build_teacher_workbook(self):
        distribution = (
            self.create_distribution()
        )

        employment = (
            distribution.staff_employment
        )
        academic_year = (
            distribution.planned_workload
            .academic_year
        )


        year_record = (
            StaffEmploymentAcademicYearFactory
            ._meta
            .model
            .objects
            .filter(
                staff_employment=employment,
                academic_year=academic_year,
            )
            .first()
        )

        self.assertIsNotNone(year_record)

        university = (
            employment.department
            .faculty.university
        )

        ExcelReportTemplateFactory(
            university=university,
            template_type=(
                ExcelReportTemplate
                .Type
                .TEACHER_WORKLOAD
            ),
            file=create_xlsx_file(
                max_columns=25,
                max_rows=6,
                placeholders={
                    "A1": "{FIO}",
                    "A2": "{year}",
                    "A3": "{st}",
                    "A4": "{degree}",
                    "B4": "{title}",
                },
            ),
        )

        result = (
            TeacherWorkloadExcelService.build(
                staff_employment_id=(
                    employment.pk
                ),
                academic_year=academic_year,
            )
        )

        workbook = load_workbook(
            result,
            data_only=False,
        )
        worksheet = workbook["Отчёт"]

        self.assertEqual(
            worksheet["A1"].value,
            employment.staff_member.full_name,
        )
        self.assertEqual(
            worksheet["A2"].value,
            str(academic_year),
        )
        self.assertEqual(
            worksheet["A5"].value,
            "Осенний семестр",
        )

        workbook.close()

    def test_missing_year_record_rejected(
        self,
    ):
        distribution = (
            self.create_distribution()
        )

        employment = (
            distribution.staff_employment
        )
        academic_year = (
            distribution.planned_workload
            .academic_year
        )

        (
            StaffEmploymentAcademicYearFactory
            ._meta
            .model
            .objects
            .filter(
                staff_employment=employment,
                academic_year=academic_year,
            )
            .delete()
        )

        with self.assertRaises(
            ReportDataError
        ):
            (
                TeacherWorkloadExcelService
                .build(
                    staff_employment_id=(
                        employment.pk
                    ),
                    academic_year=(
                        academic_year
                    ),
                )
            )


class DepartmentWorkloadExcelServiceTests(
    TestCase
):
    def test_build_empty_department_report(
        self,
    ):
        planned = PlannedWorkloadFactory()

        department = (
            planned.teaching_department
        )
        academic_year = (
            planned.academic_year
        )

        planned.delete()

        template = (
            ExcelReportTemplateFactory
            .department_template(
                university=(
                    department
                    .faculty
                    .university
                ),
            )
        )

        result = (
            DepartmentWorkloadExcelService
            .build(
                department_id=department.pk,
                academic_year=academic_year,
            )
        )

        workbook = load_workbook(
            result,
            data_only=False,
        )
        worksheet = workbook[
            template.sheet_name
        ]

        self.assertEqual(
            worksheet["A1"].value,
            department.name_ru,
        )
        self.assertEqual(
            worksheet["A2"].value,
            str(academic_year),
        )

        workbook.close()
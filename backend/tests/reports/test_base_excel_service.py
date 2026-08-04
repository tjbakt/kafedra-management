from io import BytesIO

from django.test import TestCase
from openpyxl import Workbook

from apps.reports.exceptions import (
    ReportGenerationError,
)
from apps.reports.models import (
    ExcelReportTemplate,
)
from apps.reports.services.base_excel_report import (
    BaseExcelReportService,
)
from tests.factories import (
    ExcelReportTemplateFactory,
    UniversityFactory,
    create_xlsx_file,
)


class TestExcelService(
    BaseExcelReportService
):
    template_type = (
        ExcelReportTemplate
        .Type
        .TEACHER_WORKLOAD
    )


class BaseExcelReportServiceTests(
    TestCase
):
    def test_university_template_has_priority(
        self,
    ):
        university = UniversityFactory()

        ExcelReportTemplateFactory.global_template(
            version=10,
        )

        university_template = (
            ExcelReportTemplateFactory(
                university=university,
                version=1,
            )
        )

        result = TestExcelService.get_template(
            university_id=university.pk,
        )

        self.assertEqual(
            result.pk,
            university_template.pk,
        )

    def test_falls_back_to_global_template(
        self,
    ):
        university = UniversityFactory()

        global_template = (
            ExcelReportTemplateFactory
            .global_template()
        )

        result = TestExcelService.get_template(
            university_id=university.pk,
        )

        self.assertEqual(
            result.pk,
            global_template.pk,
        )

    def test_missing_template_rejected(self):
        with self.assertRaises(
            ReportGenerationError
        ):
            TestExcelService.get_template()

    def test_service_without_type_rejected(
        self,
    ):
        with self.assertRaises(
            ReportGenerationError
        ):
            BaseExcelReportService.get_template()

    def test_missing_sheet_rejected(self):
        template = ExcelReportTemplateFactory(
            sheet_name="Несуществующий лист",
        )

        with self.assertRaises(
            ReportGenerationError
        ):
            (
                TestExcelService
                .load_template_workbook(
                    template
                )
            )

    def test_invalid_workbook_rejected(self):
        template = ExcelReportTemplateFactory(
            file=create_xlsx_file(),
        )

        template.file.save(
            "broken.xlsx",
            BytesIO(b"not an xlsx"),
            save=True,
        )

        with self.assertRaises(
            ReportGenerationError
        ):
            (
                TestExcelService
                .load_template_workbook(
                    template
                )
            )

    def test_replace_placeholders(self):
        workbook = Workbook()
        worksheet = workbook.active

        worksheet["A1"] = (
            "Преподаватель: {FIO}"
        )
        worksheet["A2"] = (
            "Учебный год: {year}"
        )

        (
            TestExcelService
            .replace_placeholders(
                worksheet,
                {
                    "{FIO}": "Иванов И.И.",
                    "{year}": "2025–2026",
                },
            )
        )

        self.assertEqual(
            worksheet["A1"].value,
            "Преподаватель: Иванов И.И.",
        )
        self.assertEqual(
            worksheet["A2"].value,
            "Учебный год: 2025–2026",
        )

        workbook.close()

    def test_set_sum_formulas(self):
        workbook = Workbook()
        worksheet = workbook.active

        (
            TestExcelService
            .set_sum_formulas(
                worksheet,
                total_row=4,
                data_start_row=2,
                data_end_row=3,
                sum_columns=(2, 3),
            )
        )

        self.assertEqual(
            worksheet["B4"].value,
            "=SUM(B2:B3)",
        )
        self.assertEqual(
            worksheet["C4"].value,
            "=SUM(C2:C3)",
        )

        workbook.close()

    def test_save_to_bytes(self):
        workbook = Workbook()
        workbook.active["A1"] = "Test"

        result = (
            TestExcelService.save_to_bytes(
                workbook
            )
        )

        self.assertIsInstance(
            result,
            BytesIO,
        )
        self.assertGreater(
            len(result.getvalue()),
            0,
        )